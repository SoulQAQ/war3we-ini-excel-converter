#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Excel 转 INI 工具
将 Excel 工作簿转换回 Warcraft III table 目录常用的 INI 文件。
"""

import os
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook


MULTILINE_ELEMENT_SEPARATOR = '----'
EMPTY_TEXT_PLACEHOLDER = '@empty'
METADATA_SHEET_NAME = '__ini_meta'
SOURCE_SHEET_NAME = '__ini_sources'

DEFAULT_SHEET_FILENAME_MAP = {
    '技能': 'ability.ini',
    '魔法效果': 'buff.ini',
    '物品': 'item.ini',
    '单位': 'unit.ini',
    '科技': 'upgrade.ini',
}

PARENT_HEADER_LABELS = {'模板 id', '模板ID', '模板 ID', '父级 id', '父级ID', '父级 ID', 'parent'}
ID_HEADER_LABELS = {'物体 id', '物体ID', '物体 ID', 'id', 'object id'}
NUMERIC_PATTERN = re.compile(r'^[+-]?(?:\d+(?:\.\d*)?|\.\d+)$')
CALC_PATTERN = re.compile(r'^calc@([+-]?(?:\d+(?:\.\d*)?|\.\d+))([+-])((?:\d+(?:\.\d*)?|\.\d+))$')
INVALID_FILENAME_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')

LEVEL_ARRAY_PROPERTIES = {
    'area',
    'buffid',
    'cast',
    'cool',
    'cost',
    'dur',
    'efctid',
    'effectid',
    'herodur',
    'hotkey',
    'order',
    'orderoff',
    'orderon',
    'requires',
    'requiresamount',
    'researchhotkey',
    'researchtip',
    'researchubertip',
    'rng',
    'targs',
    'targets',
    'tip',
    'unitid',
    'untip',
    'unubertip',
    'ubertip',
}


def normalize_cell_text(value):
    if value is None:
        return ''
    return str(value).replace('\r\n', '\n').replace('\r', '\n')


def get_header_text(ws, row, col):
    return normalize_cell_text(ws.cell(row=row, column=col).value).strip()


def normalize_header(text):
    return re.sub(r'\s+', ' ', str(text or '')).strip()


def normalize_header_key(text):
    return normalize_header(text).lower()


def is_numeric_token(text):
    return bool(NUMERIC_PATTERN.fullmatch(str(text).strip()))


def is_level_array_property(prop_name):
    normalized = str(prop_name or '').strip().lower()
    if normalized in LEVEL_ARRAY_PROPERTIES:
        return True
    return bool(re.fullmatch(r'data[a-z][a-z0-9_]*', normalized))


def format_decimal(value, force_decimal=False):
    text = format(value, 'f')
    if '.' in text:
        text = text.rstrip('0').rstrip('.')
    if text in {'', '-0'}:
        text = '0'
    if force_decimal and '.' not in text:
        text += '.0'
    return text


def parse_positive_int(value):
    text = normalize_cell_text(value).strip()
    if not text:
        return 0

    try:
        parsed = Decimal(text)
    except InvalidOperation:
        return 0

    if parsed <= 0:
        return 0
    return int(parsed)


def parse_metadata_count(value):
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def split_top_level_csv(text):
    """按顶层逗号切分，忽略引号、尖括号和 Lua 长字符串中的逗号。"""
    if not isinstance(text, str):
        return []

    parts = []
    current = []
    in_quotes = False
    angle_depth = 0
    lua_block_depth = 0
    i = 0

    while i < len(text):
        if text.startswith('[=[', i):
            lua_block_depth += 1
            current.append('[=[')
            i += 3
            continue
        if lua_block_depth > 0 and text.startswith(']=]', i):
            lua_block_depth -= 1
            current.append(']=]')
            i += 3
            continue

        char = text[i]
        if lua_block_depth == 0:
            if char == '"':
                in_quotes = not in_quotes
            elif not in_quotes:
                if char == '<':
                    angle_depth += 1
                elif char == '>' and angle_depth > 0:
                    angle_depth -= 1
                elif char == ',' and angle_depth == 0:
                    parts.append(''.join(current).strip())
                    current = []
                    i += 1
                    continue

        current.append(char)
        i += 1

    trailing = ''.join(current).strip()
    if trailing:
        parts.append(trailing)
    elif parts:
        parts.append('')
    return parts


def split_multiline_elements(text):
    """将 Excel 中使用 ---- 分隔的多等级内容拆成元素。"""
    lines = text.split('\n')
    if not any(line.strip() == MULTILINE_ELEMENT_SEPARATOR for line in lines):
        return None

    parts = []
    current_part = []
    for line in lines:
        if line.strip() == MULTILINE_ELEMENT_SEPARATOR:
            parts.append('\n'.join(current_part).strip('\n'))
            current_part = []
        else:
            current_part.append(line)
    parts.append('\n'.join(current_part).strip('\n'))

    return [part if part != '' else EMPTY_TEXT_PLACEHOLDER for part in parts]


def expand_calc_formula(text, level_count):
    """把 calc@A+B / calc@A-B 展开为等级数组。"""
    if level_count <= 0:
        return None

    match = CALC_PATTERN.fullmatch(text.strip())
    if not match:
        return None

    base_text, symbol, step_text = match.groups()
    try:
        base = Decimal(base_text)
        step = Decimal(step_text)
    except InvalidOperation:
        return None

    if symbol == '-':
        step = -step

    force_decimal = '.' in base_text or '.' in step_text
    return [format_decimal(base + step * index, force_decimal) for index in range(level_count)]


def encode_ini_scalar(value):
    text = normalize_cell_text(value)
    stripped = text.strip()

    if stripped == EMPTY_TEXT_PLACEHOLDER:
        return '""'
    if stripped == '':
        return '""'
    if stripped.startswith('[=[') and stripped.endswith(']=]'):
        return stripped
    if stripped.startswith('"') and stripped.endswith('"') and len(stripped) >= 2:
        return stripped
    if '\n' in text:
        return f'[=[\n{text.strip("\n")}\n]=]'
    if is_numeric_token(stripped):
        return stripped

    return f'"{stripped.replace(chr(34), chr(92) + chr(34))}"'


def render_ini_array(elements):
    normalized = [EMPTY_TEXT_PLACEHOLDER if element is None else normalize_cell_text(element).strip('\n') for element in elements]
    encoded = [encode_ini_scalar(element) for element in normalized]

    if not encoded:
        return '{}'

    inline_value = '{' + ', '.join(encoded) + '}'
    has_long_string = any('\n' in item for item in encoded)
    has_text = any(item.startswith('"') or item.startswith('[=[') for item in encoded)

    if not has_long_string and not has_text and len(inline_value) <= 100:
        return inline_value

    lines = ['{']
    for item in encoded:
        if '\n' in item:
            item_lines = item.split('\n')
            lines.append(item_lines[0])
            lines.extend(item_lines[1:])
            lines[-1] += ','
        else:
            lines.append(f'{item},')
    lines.append('}')
    return '\n'.join(lines)


def get_array_target_count(value_meta, level_count):
    if not value_meta:
        return level_count

    metadata_count = parse_metadata_count(value_meta.get('element_count'))
    if metadata_count > 0:
        return metadata_count
    return level_count


def encode_ini_value(prop_name, value, level_count=0, value_meta=None):
    """将 Excel 单元格中的值编码为 INI 中使用的文本格式。"""
    text = normalize_cell_text(value)
    stripped = text.strip()

    if stripped == '':
        return ''

    value_kind = (value_meta or {}).get('value_kind', '')
    array_target_count = get_array_target_count(value_meta, level_count)

    calc_elements = expand_calc_formula(stripped, array_target_count)
    if calc_elements:
        return render_ini_array(calc_elements)

    multiline_elements = split_multiline_elements(text)
    if multiline_elements is not None:
        return render_ini_array(multiline_elements)

    csv_elements = split_top_level_csv(stripped)
    if (
        array_target_count > 1
        and is_level_array_property(prop_name)
        and len(csv_elements) == array_target_count
    ):
        return render_ini_array(csv_elements)

    if value_kind == 'array':
        target_count = max(array_target_count, 1)
        return render_ini_array([stripped] * target_count)

    return encode_ini_scalar(text)


def load_workbook_metadata(workbook):
    """读取隐藏元数据表，记录每个属性原本是标量还是数组。"""
    if METADATA_SHEET_NAME not in workbook.sheetnames:
        return {}

    ws = workbook[METADATA_SHEET_NAME]
    metadata = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        sheet_name, object_id, property_name, value_kind, element_count = (list(row) + [None] * 5)[:5]
        sheet_name = normalize_cell_text(sheet_name).strip()
        object_id = normalize_cell_text(object_id).strip()
        property_name = normalize_cell_text(property_name).strip()
        if not sheet_name or not object_id or not property_name:
            continue

        metadata[(sheet_name, object_id, property_name)] = {
            'value_kind': normalize_cell_text(value_kind).strip(),
            'element_count': parse_metadata_count(element_count),
        }
    return metadata


def load_workbook_sources(workbook):
    """读取隐藏源文件表，定位每个工作表原本对应的 INI 文件。"""
    if SOURCE_SHEET_NAME not in workbook.sheetnames:
        return {}

    ws = workbook[SOURCE_SHEET_NAME]
    sources = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        sheet_name, ini_filename, source_path = (list(row) + [None] * 3)[:3]
        sheet_name = normalize_cell_text(sheet_name).strip()
        ini_filename = normalize_cell_text(ini_filename).strip()
        source_path = normalize_cell_text(source_path).strip()
        if not sheet_name:
            continue

        sources[sheet_name] = {
            'ini_filename': ini_filename,
            'source_path': source_path,
        }
    return sources


def detect_sheet_layout(ws):
    """识别当前导出表和旧表的 ID、父模板和属性列。"""
    max_column = ws.max_column
    id_col = 1
    parent_col = None
    property_columns = []

    for col in range(1, max_column + 1):
        label = get_header_text(ws, 1, col)
        prop_name = get_header_text(ws, 2, col)
        label_key = normalize_header_key(label)
        prop_key = normalize_header_key(prop_name)

        if col == 1 or label_key in {item.lower() for item in ID_HEADER_LABELS}:
            id_col = col
            continue

        if prop_key == '_parent' or label in PARENT_HEADER_LABELS or label_key in {item.lower() for item in PARENT_HEADER_LABELS}:
            parent_col = col
            continue

        if prop_name:
            property_columns.append((col, prop_name, label))

    if parent_col is None:
        second_prop = get_header_text(ws, 2, 2)
        if normalize_header_key(second_prop) not in {'name', 'levels'}:
            parent_col = 2
            property_columns = [item for item in property_columns if item[0] != 2]

    return id_col, parent_col, property_columns


def collect_row_values(ws, row, property_columns):
    row_values = {}
    for col, prop_name, _comment in property_columns:
        cell_value = ws.cell(row=row, column=col).value
        if cell_value is None or normalize_cell_text(cell_value).strip() == '':
            continue
        row_values[prop_name] = cell_value
    return row_values


def get_row_level_count(row_values):
    for key in ('levels', 'maxlevel'):
        for prop_name, value in row_values.items():
            if prop_name.lower() == key:
                level_count = parse_positive_int(value)
                if level_count > 0:
                    return level_count
    return 0


def append_ini_assignment(lines, prop_name, encoded_value):
    if '\n' not in encoded_value:
        lines.append(f'{prop_name} = {encoded_value}')
        return

    encoded_lines = encoded_value.split('\n')
    lines.append(f'{prop_name} = {encoded_lines[0]}')
    lines.extend(encoded_lines[1:])


def build_ini_lines_from_sheet(ws, workbook_metadata=None):
    """将单个工作表转换为 INI 文本行。"""
    lines = []
    max_row = ws.max_row
    id_col, parent_col, property_columns = detect_sheet_layout(ws)
    workbook_metadata = workbook_metadata or {}

    for row in range(3, max_row + 1):
        object_id = ws.cell(row=row, column=id_col).value
        if object_id is None or normalize_cell_text(object_id).strip() == '':
            continue

        object_id = normalize_cell_text(object_id).strip()
        parent_id = ''
        if parent_col:
            parent_id = normalize_cell_text(ws.cell(row=row, column=parent_col).value).strip()

        row_values = collect_row_values(ws, row, property_columns)
        level_count = get_row_level_count(row_values)

        lines.append(f'[{object_id}]')
        if parent_id:
            lines.append(f'_parent = {encode_ini_scalar(parent_id)}')

        for col, prop_name, comment in property_columns:
            cell_value = ws.cell(row=row, column=col).value
            if cell_value is None or normalize_cell_text(cell_value).strip() == '':
                continue

            value_meta = workbook_metadata.get((ws.title, object_id, prop_name), {})
            encoded_value = encode_ini_value(prop_name, cell_value, level_count, value_meta)
            if encoded_value == '':
                continue

            if comment:
                lines.append(f'-- {comment}')
            append_ini_assignment(lines, prop_name, encoded_value)

        lines.append('')

    return lines


def build_sheet_filename_map(ini_names=None):
    sheet_filename_map = dict(DEFAULT_SHEET_FILENAME_MAP)
    for filename, sheet_name in (ini_names or {}).items():
        cleaned_filename = os.path.basename(str(filename).strip())
        cleaned_sheet_name = str(sheet_name).strip()
        if cleaned_filename and cleaned_sheet_name:
            sheet_filename_map[cleaned_sheet_name] = cleaned_filename
    return sheet_filename_map


def sanitize_ini_filename(sheet_name):
    cleaned = INVALID_FILENAME_CHARS.sub('_', str(sheet_name).strip()).strip('. ')
    if not cleaned:
        cleaned = 'sheet'
    if not cleaned.lower().endswith('.ini'):
        cleaned += '.ini'
    return cleaned


def write_ini_file(path, lines):
    parent_dir = os.path.dirname(path)
    if parent_dir and not os.path.exists(parent_dir):
        os.makedirs(parent_dir)

    with open(path, 'w', encoding='utf-8', newline='\n') as file:
        file.write('\n'.join(lines).rstrip() + '\n')


def excel_to_ini(excel_path, output_path=None, ini_names=None, prefer_source_paths=False):
    """
    将 Excel 文件转换为 INI 文件或 table 目录。

    Args:
        excel_path: Excel 文件路径
        output_path: 输出的 INI 文件路径或文件夹路径
        ini_names: INI 文件名到中文 sheet 名称的映射字典
        prefer_source_paths: 为 True 时，优先回写 Excel 元数据记录的原始 INI 路径
    """
    excel_file_path = Path(excel_path).resolve()
    workbook = load_workbook(excel_file_path)
    workbook_metadata = load_workbook_metadata(workbook)
    workbook_sources = load_workbook_sources(workbook)

    if prefer_source_paths and workbook_sources:
        written_files = []
        fallback_output_dir = Path(output_path) if output_path else None
        sheet_filename_map = build_sheet_filename_map(ini_names)

        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            if ws.sheet_state == 'hidden' or sheet_name.startswith('__'):
                continue

            lines = build_ini_lines_from_sheet(ws, workbook_metadata)
            source_info = workbook_sources.get(sheet_name, {})
            source_path = source_info.get('source_path') or ''
            if source_path:
                source_candidate = Path(source_path)
                if source_candidate.is_absolute():
                    ini_path = source_candidate
                else:
                    ini_path = (excel_file_path.parent / source_candidate).resolve()
            elif fallback_output_dir:
                filename = source_info.get('ini_filename') or sheet_filename_map.get(sheet_name, sanitize_ini_filename(sheet_name))
                ini_path = fallback_output_dir / filename
            else:
                raise ValueError(f'工作表 {sheet_name} 缺少原始 INI 路径，且未指定备用输出目录。')

            write_ini_file(str(ini_path), lines)
            written_files.append(str(ini_path))
            print(f'INI 文件已回写：{ini_path}')

        return written_files

    if not output_path:
        raise ValueError('缺少输出路径。此 Excel 没有源文件路径元数据，无法直接回写原位置。')

    target_path = Path(output_path)

    if target_path.suffix.lower() == '.ini':
        ws = workbook[workbook.sheetnames[0]]
        lines = build_ini_lines_from_sheet(ws, workbook_metadata)
        write_ini_file(str(target_path), lines)
        print(f'INI 文件已创建：{target_path}')
        return [str(target_path)]

    output_dir = target_path
    output_dir.mkdir(parents=True, exist_ok=True)
    sheet_filename_map = build_sheet_filename_map(ini_names)
    written_files = []

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        if ws.sheet_state == 'hidden' or sheet_name.startswith('__'):
            continue

        lines = build_ini_lines_from_sheet(ws, workbook_metadata)
        filename = sheet_filename_map.get(sheet_name, sanitize_ini_filename(sheet_name))
        ini_path = output_dir / filename
        write_ini_file(str(ini_path), lines)
        written_files.append(str(ini_path))
        print(f'INI 文件已创建：{ini_path}')

    return written_files


if __name__ == "__main__":
    excel_to_ini("./test.xlsx", "./output")
