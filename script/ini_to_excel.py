#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
INI 转 Excel 工具
将 INI 文件转换为 Excel 文件
"""

import os
import sys
import re
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def get_ini_files(folder_path):
    """
    获取文件夹内所有 INI 文件
    
    Args:
        folder_path: 文件夹路径
        
    Returns:
        list: INI 文件路径列表
    """
    ini_files = []
    if os.path.isdir(folder_path):
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                if file.lower().endswith('.ini'):
                    ini_files.append(os.path.join(root, file))
    elif os.path.isfile(folder_path) and folder_path.lower().endswith('.ini'):
        ini_files.append(folder_path)
    return ini_files


def decode_ini_value(value):
    """将 INI 原始值解码为适合 Excel 展示的纯文本。"""
    if not isinstance(value, str):
        return value

    stripped = value.strip()

    if stripped.startswith('{') and stripped.endswith('}'):
        inner = stripped[1:-1].strip()
        if inner:
            matches = re.findall(r'\[=\[(.*?)\]=\]', inner, flags=re.DOTALL)
            if matches:
                return '\n\n'.join(match.strip('\n\r') for match in matches)
        return inner

    if stripped.startswith('[=[') and stripped.endswith(']=]'):
        return stripped[3:-3]

    if value.startswith('"') and value.endswith('"') and '\n' not in value:
        return value[1:-1]

    return value


def parse_ini_file(file_path):
    """
    解析 INI 文件，提取物体 ID、模板 ID、注释和属性。

    关键规则：
    - 属性行必须是：<属性名> = <值>
      属性名仅允许字母/数字/下划线（例如 Ubertip、AGI、_parent）
    - 多行文本以 [=[ 开始，以 ]=] 结束，期间所有内容都属于同一个属性值
    - 多行文本中的任何 "=" 都不参与属性解析
    - 写入 Excel 时：
      * 头部单元格只显示“属性说明”，不包含 [=[ 或 ]=]
      * 数据单元格只显示纯内容，便于编辑与回写
    """
    objects = []
    current_object = None
    current_comment = None

    in_multiline = False
    multiline_key = None
    multiline_value_lines = []
    multiline_comment = None

    section_pattern = re.compile(r'^\s*\[([^\]]+)\]\s*$')
    property_pattern = re.compile(r'^\s*([A-Za-z_][A-Za-z0-9_]*)\s*=\s*(.*)$')

    def append_property(obj, key, value, comment):
        if obj is None:
            return
        if key == '_parent':
            obj['parent'] = value.strip('"')
        else:
            obj['properties'].append({
                'name': key,
                'value': value,
                'comment': comment
            })

    def parse_braced_multiline_value(initial_text, line_iterator):
        """
        解析形如 { [=[...]=], [=[...]=], ... } 的集合多行文本，保留完整原始文本内容。
        返回值不包含属性名左侧，仅包含从 { 开始到 } 结束的值文本。
        """
        collected_lines = [initial_text]
        bracket_depth = initial_text.count('{') - initial_text.count('}')
        in_lua_block = False

        while bracket_depth > 0:
            try:
                next_raw_line = next(line_iterator)
            except StopIteration:
                break

            next_line = next_raw_line.rstrip('\n\r')
            collected_lines.append(next_line)

            cursor = 0
            while cursor < len(next_line):
                if not in_lua_block and next_line.startswith('[=[', cursor):
                    in_lua_block = True
                    cursor += 3
                    continue
                if in_lua_block and next_line.startswith(']=]', cursor):
                    in_lua_block = False
                    cursor += 3
                    continue
                if not in_lua_block:
                    char = next_line[cursor]
                    if char == '{':
                        bracket_depth += 1
                    elif char == '}':
                        bracket_depth -= 1
                cursor += 1

        return '\n'.join(collected_lines)

    with open(file_path, 'r', encoding='utf-8') as f:
        line_iterator = iter(f)
        for raw_line in line_iterator:
            line = raw_line.rstrip('\n\r')
            stripped = line.strip()

            # 1) 若在多行文本块中，只查找结束标记 ]=]
            if in_multiline:
                end_pos = line.find(']=]')
                if end_pos != -1:
                    tail = line[:end_pos]
                    if tail:
                        multiline_value_lines.append(tail)
                    full_value = '\n'.join(multiline_value_lines)
                    full_value = full_value.lstrip('\n\r')
                    append_property(current_object, multiline_key, full_value, multiline_comment)

                    in_multiline = False
                    multiline_key = None
                    multiline_value_lines = []
                    multiline_comment = None
                else:
                    multiline_value_lines.append(line)
                continue

            # 2) 段落头 [XXXX]
            section_match = section_pattern.match(line)
            if section_match:
                if current_object is not None:
                    objects.append(current_object)
                current_object = {
                    'id': section_match.group(1),
                    'parent': '',
                    'properties': []
                }
                current_comment = None
                continue

            if current_object is None:
                continue

            # 3) 注释行
            if stripped.startswith('--'):
                current_comment = stripped[2:].strip()
                continue

            # 4) 空行
            if stripped == '':
                current_comment = None
                continue

            # 5) 属性行（严格匹配属性名）
            prop_match = property_pattern.match(line)
            if not prop_match:
                # 非法属性行直接忽略，防止把多行文本尾部误识别为表头
                continue

            key = prop_match.group(1)
            value = prop_match.group(2).strip()

            # 6) 多行/集合文本开始
            if value.startswith('{'):
                # ability.ini 中存在 { [=[...]=], ... } 这类 Lua 风格集合文本。
                # 旧逻辑会把起始 { 视为纯结构符号并丢弃，导致导出 Excel 为空。
                # 这里直接按完整块保留原始值，直到匹配到对应的 }。
                braced_value = parse_braced_multiline_value(value, line_iterator)
                append_property(current_object, key, braced_value, current_comment)
                current_comment = None
                continue

            # 7) 多行文本开始（值以 [=[ 开头）
            if value.startswith('[=['):
                after_start = value[3:]
                end_pos = after_start.find(']=]')

                if end_pos != -1:
                    # 单行写完：[=[xxx]=]
                    content = after_start[:end_pos]
                    append_property(current_object, key, content, current_comment)
                    current_comment = None
                else:
                    # 多行开始
                    in_multiline = True
                    multiline_key = key
                    multiline_comment = current_comment
                    multiline_value_lines = [after_start] if after_start else []
                    current_comment = None
                continue

            # 8) 普通单行属性
            append_property(current_object, key, value, current_comment)
            current_comment = None

    # 容错：文件结尾仍处于多行文本状态，按已收集内容写入
    if in_multiline and current_object is not None and multiline_key:
        full_value = '\n'.join(multiline_value_lines)
        full_value = full_value.lstrip('\n\r')
        append_property(current_object, multiline_key, full_value, multiline_comment)

    if current_object is not None:
        objects.append(current_object)

    return objects


def auto_size_column(ws, col, min_width=10, max_width=100):
    """
    自动调整列宽
    
    Args:
        ws: 工作表
        col: 列号
        min_width: 最小宽度
        max_width: 最大宽度
    """
    max_len = min_width
    for row in ws.iter_rows(min_col=col, max_col=col):
        cell = row[0]
        if cell.value is not None:
            cell_len = len(str(cell.value))
            if cell_len > max_len:
                max_len = cell_len
    
    # 限制最大宽度
    if max_len > max_width:
        max_len = max_width
    
    ws.column_dimensions[get_column_letter(col)].width = max_len


def get_unique_filename(filepath):
    """
    如果文件已存在，在文件名后添加_1、_2 等后缀
    
    Args:
        filepath: 原始文件路径
        
    Returns:
        str: 不冲突的文件路径
    """
    if not os.path.exists(filepath):
        return filepath
    
    directory = os.path.dirname(filepath)
    filename = os.path.basename(filepath)
    name, ext = os.path.splitext(filename)
    
    # 如果名字已经以 _数字 结尾，去掉它以便重新计数
    match = re.match(r'^(.+?)_(\d+)$', name)
    if match:
        name = match.group(1)
    
    counter = 1
    while True:
        new_filename = f"{name}_{counter}{ext}"
        new_filepath = os.path.join(directory, new_filename)
        if not os.path.exists(new_filepath):
            return new_filepath
        counter += 1


def create_excel_with_sheets(ini_folder, output_path, ini_names=None):
    """
    创建 Excel 文件，每个 INI 文件作为一个 sheet
    
    Args:
        ini_folder: INI 文件所在文件夹路径
        output_path: 输出的 Excel 文件路径
        ini_names: INI 文件名到中文名称的映射字典
    """
    wb = Workbook()
    
    ini_files = get_ini_files(ini_folder)
    
    if not ini_files:
        print(f"未在 {ini_folder} 中找到 INI 文件")
        return
    
    print(f"找到 {len(ini_files)} 个 INI 文件")
    
    for idx, ini_file in enumerate(ini_files):
        # 使用文件名（不含扩展名）作为 sheet 名
        filename = os.path.basename(ini_file)
        file_key = filename.lower()
        
        # 如果有中文名称映射，使用中文名称
        if ini_names and file_key in ini_names:
            sheet_name = ini_names[file_key]
        else:
            sheet_name = os.path.splitext(filename)[0]
        
        # Excel sheet 名长度限制为 31 字符
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]
        
        if idx == 0:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(title=sheet_name)
        
        # 解析 INI 文件
        objects = parse_ini_file(ini_file)
        
        # 写入数据
        # 第一行：物体 ID 和模板 ID
        ws.cell(row=1, column=1, value='物体 ID')
        ws.cell(row=2, column=1, value='')
        ws.cell(row=1, column=2, value='模板 ID')
        ws.cell(row=2, column=2, value='')

        yahei_font = Font(name='Microsoft YaHei')
        
        # 收集所有属性名和注释
        all_props = []
        prop_comments = {}
        for obj in objects:
            for prop in obj['properties']:
                if prop['name'] not in prop_comments:
                    all_props.append(prop['name'])
                if prop['comment']:
                    prop_comments[prop['name']] = prop['comment']
        
        # 从第 3 列开始写入属性
        col_offset = 3
        for col_idx, prop_name in enumerate(all_props):
            col = col_offset + col_idx
            # 第一行表头：注释
            comment = prop_comments.get(prop_name, '')
            ws.cell(row=1, column=col, value=comment)
            # 第二行表头：属性名
            ws.cell(row=2, column=col, value=prop_name)
        
        # 写入物体数据
        for row_idx, obj in enumerate(objects, start=3):
            ws.cell(row=row_idx, column=1, value=obj['id'])
            ws.cell(row=row_idx, column=2, value=obj['parent'])
            
            # 写入属性值
            prop_values = {p['name']: p['value'] for p in obj['properties']}
            for col_idx, prop_name in enumerate(all_props):
                col = col_offset + col_idx
                value = decode_ini_value(prop_values.get(prop_name, ''))
                ws.cell(row=row_idx, column=col, value=value)
        
        for row in ws.iter_rows():
            for cell in row:
                cell.font = yahei_font

        ws.freeze_panes = 'C3'

        # 自动调整列宽
        # 调整物体 ID 和模板 ID 列
        auto_size_column(ws, 1, min_width=10, max_width=20)
        auto_size_column(ws, 2, min_width=10, max_width=20)
        
        # 调整属性列
        for col_idx in range(len(all_props)):
            col = col_offset + col_idx
            auto_size_column(ws, col, min_width=15, max_width=80)
    
    # 保存 Excel 文件
    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    wb.save(output_path)
    print(f"Excel 文件已创建：{output_path}")
    print(f"共创建 {len(ini_files)} 个 sheet")


def ini_to_excel(config_path, output_path, ini_names=None):
    """
    将 INI 文件转换为 Excel 文件
    
    Args:
        config_path: INI 文件路径或文件夹路径
        output_path: 输出的 Excel 文件路径
        ini_names: INI 文件名到中文名称的映射字典
    """
    create_excel_with_sheets(config_path, output_path, ini_names)


if __name__ == "__main__":
    # 测试调用
    ini_to_excel("./test", "./output.xlsx")
